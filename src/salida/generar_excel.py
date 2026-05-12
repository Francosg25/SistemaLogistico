"""
Generador principal del Excel de salida.

Orquesta la creación de todas las hojas, usando fórmulas Excel reales
(no valores hard-coded) para que el usuario pueda auditar los cálculos.
"""
from io import BytesIO
from pathlib import Path
from datetime import datetime
from typing import Optional, Union
from openpyxl import Workbook

from src.salida.hoja_summary import construir_hoja_summary
from src.salida.hoja_outbound import construir_hoja_outbound
from src.salida.hoja_sea import construir_hoja_sea
from src.salida.hoja_land import construir_hoja_land
from src.salida.hoja_validaciones import construir_hoja_validaciones
from src.salida.hoja_reglas import construir_hoja_reglas
from src.utils.logger import configurar_logger

logger = configurar_logger("generar_excel")


def generar_excel_completo(
    resultado_outbound,
    resultado_sea,
    resultado_land,
    resultado_summary,
    reporte_validacion=None,
    costos: dict = None,
    ruta_salida: Optional[Union[str, Path]] = None,
) -> BytesIO:
    """
    Genera el Excel de salida completo con todas las hojas.
    
    Args:
        resultado_outbound: ResultadoOutbound del Bloque 3
        resultado_sea: ResultadoSea del Bloque 4
        resultado_land: ResultadoLand del Bloque 5
        resultado_summary: ResultadoSummary del Bloque 6
        reporte_validacion: ReporteValidacion del Bloque 8 (opcional)
        costos: dict con costos fijos {'outbound': 1500, 'sea': 2500, 'land': 1200}
        ruta_salida: Si se proporciona, guarda el archivo en disco
    
    Returns:
        BytesIO con el archivo Excel (para descarga en Streamlit)
    """
    logger.info("=" * 60)
    logger.info("💾 INICIANDO GENERACIÓN DE EXCEL DE SALIDA")
    logger.info("=" * 60)
    
    costos = costos or {"outbound": 1500, "sea": 2500, "land": 1200}
    
    # ─────────────────────────────────────────────────────────
    # 1. CREAR WORKBOOK
    # ─────────────────────────────────────────────────────────
    wb = Workbook()
    # Eliminar la hoja default
    wb.remove(wb.active)
    
    # ─────────────────────────────────────────────────────────
    # 2. CREAR HOJAS EN EL ORDEN CORRECTO
    # ─────────────────────────────────────────────────────────
    # ORDEN IMPORTANTE: Las hojas que son REFERENCIADAS por Summary
    # deben crearse ANTES para que las fórmulas funcionen al abrir.
    
    # Hoja Outbound (referenciada por Summary!Outbound!E:E, G:G)
    logger.info("📤 Construyendo hoja Outbound...")
    ws_outbound = wb.create_sheet("Outbound")
    construir_hoja_outbound(ws_outbound, resultado_outbound, costos["outbound"])
    
    # Hoja Sea (referenciada por Summary!Sea!E:E, H:H)
    logger.info("🚢 Construyendo hoja Sea...")
    ws_sea = wb.create_sheet("Sea")
    construir_hoja_sea(ws_sea, resultado_sea, costos["sea"])
    
    # Hoja Land (referenciada por Summary!Land!E:E, G:G)
    logger.info("🚚 Construyendo hoja Land...")
    ws_land = wb.create_sheet("Land")
    construir_hoja_land(ws_land, resultado_land, costos["land"])
    
    # Hoja Summary (USA fórmulas que referencian las anteriores)
    logger.info("📊 Construyendo hoja Summary...")
    ws_summary = wb.create_sheet("Summary", 0)  # Insertar al principio
    construir_hoja_summary(ws_summary, resultado_summary.bus_orden)
    
    # Hoja Validaciones (opcional)
    if reporte_validacion is not None:
        logger.info("✅ Construyendo hoja Validaciones...")
        ws_val = wb.create_sheet("Validaciones")
        construir_hoja_validaciones(ws_val, reporte_validacion)
    
    # Hoja REGLAS_PROCESO (documentación)
    logger.info("📚 Construyendo hoja REGLAS_PROCESO...")
    ws_reglas = wb.create_sheet("REGLAS_PROCESO")
    construir_hoja_reglas(ws_reglas)
    
    # ─────────────────────────────────────────────────────────
    # 3. METADATA DEL WORKBOOK
    # ─────────────────────────────────────────────────────────
    wb.properties.title = "Consolidación Logística"
    wb.properties.creator = "Software Consolidación v1.0"
    wb.properties.created = datetime.now()
    wb.properties.description = (
        "Reporte consolidado de Outbound, Sea y Land con Summary por BU. "
        "Generado automáticamente."
    )
    
    # ─────────────────────────────────────────────────────────
    # 4. GUARDAR EN MEMORIA (BytesIO) Y OPCIONALMENTE EN DISCO
    # ─────────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    if ruta_salida:
        ruta = Path(ruta_salida)
        ruta.parent.mkdir(parents=True, exist_ok=True)
        with open(ruta, "wb") as f:
            f.write(buffer.getvalue())
        logger.info(f"💾 Excel guardado en: {ruta}")
        buffer.seek(0)
    
    logger.info("─" * 60)
    logger.info(f"✅ Excel generado exitosamente")
    logger.info(f"📑 Hojas creadas: {wb.sheetnames}")
    logger.info("=" * 60)
    
    return buffer


# ============================================================
# FUNCIONES AUXILIARES PARA EXPORTACIÓN INDIVIDUAL
# ============================================================
def generar_excel_solo_sea(resultado_sea, costo: float = 2500) -> BytesIO:
    """Genera un Excel con SOLO la hoja Sea (para descarga independiente)."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Sea")
    construir_hoja_sea(ws, resultado_sea, costo)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_solo_land(resultado_land, costo: float = 1200) -> BytesIO:
    """Genera un Excel con SOLO la hoja Land."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Land")
    construir_hoja_land(ws, resultado_land, costo)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_solo_outbound(resultado_outbound, costo: float = 1500) -> BytesIO:
    """Genera un Excel con SOLO la hoja Outbound."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Outbound")
    construir_hoja_outbound(ws, resultado_outbound, costo)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer