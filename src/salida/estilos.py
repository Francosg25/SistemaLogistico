"""
Estilos visuales centralizados para el Excel de salida.
Todos los colores, fuentes y formatos se definen aquí para mantener consistencia.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter


# ============================================================
# PALETA DE COLORES (corporativa)
# ============================================================
COLORES = {
    "azul_primario":    "1F4E78",   # Azul oscuro - encabezados principales
    "azul_secundario":  "2E75B6",   # Azul medio - subtítulos
    "azul_claro":       "DDEBF7",   # Azul muy claro - filas alternadas
    "verde_ok":         "C6EFCE",   # Verde - validación OK
    "amarillo_warn":    "FFEB9C",   # Amarillo - warnings
    "rojo_error":       "FFC7CE",   # Rojo claro - errores
    "gris_claro":       "F2F2F2",   # Gris - filas no destacadas
    "blanco":           "FFFFFF",
    "negro":            "000000",
    "naranja_capex":    "FCE4D6",   # CAPEX especial
    "morado_excluido":  "E4DFEC",   # Capex/MCS excluidos
}


# ============================================================
# FUENTES
# ============================================================
FUENTES = {
    "titulo": Font(name="Calibri", size=16, bold=True, color=COLORES["blanco"]),
    "subtitulo": Font(name="Calibri", size=12, bold=True, color=COLORES["blanco"]),
    "encabezado": Font(name="Calibri", size=11, bold=True, color=COLORES["blanco"]),
    "encabezado_oscuro": Font(name="Calibri", size=11, bold=True, color=COLORES["negro"]),
    "normal": Font(name="Calibri", size=10, color=COLORES["negro"]),
    "negrita": Font(name="Calibri", size=10, bold=True, color=COLORES["negro"]),
    "total": Font(name="Calibri", size=11, bold=True, color=COLORES["blanco"]),
    "caption": Font(name="Calibri", size=9, italic=True, color="595959"),
}


# ============================================================
# RELLENOS
# ============================================================
RELLENOS = {
    "titulo":           PatternFill("solid", fgColor=COLORES["azul_primario"]),
    "subtitulo":        PatternFill("solid", fgColor=COLORES["azul_secundario"]),
    "encabezado":       PatternFill("solid", fgColor=COLORES["azul_primario"]),
    "fila_alterna":     PatternFill("solid", fgColor=COLORES["azul_claro"]),
    "fila_normal":      PatternFill("solid", fgColor=COLORES["blanco"]),
    "total":            PatternFill("solid", fgColor=COLORES["azul_secundario"]),
    "ok":               PatternFill("solid", fgColor=COLORES["verde_ok"]),
    "warning":          PatternFill("solid", fgColor=COLORES["amarillo_warn"]),
    "error":            PatternFill("solid", fgColor=COLORES["rojo_error"]),
    "capex":            PatternFill("solid", fgColor=COLORES["naranja_capex"]),
    "excluido":         PatternFill("solid", fgColor=COLORES["morado_excluido"]),
}


# ============================================================
# ALINEACIONES
# ============================================================
ALINEACIONES = {
    "centro":   Alignment(horizontal="center", vertical="center", wrap_text=True),
    "izquierda": Alignment(horizontal="left", vertical="center"),
    "derecha":  Alignment(horizontal="right", vertical="center"),
}


# ============================================================
# BORDES
# ============================================================
_borde_fino = Side(style="thin", color="BFBFBF")
_borde_medio = Side(style="medium", color="808080")

BORDES = {
    "fino": Border(left=_borde_fino, right=_borde_fino, top=_borde_fino, bottom=_borde_fino),
    "medio": Border(left=_borde_medio, right=_borde_medio, top=_borde_medio, bottom=_borde_medio),
    "ninguno": Border(),
}


# ============================================================
# FORMATOS NUMÉRICOS
# ============================================================
FORMATOS = {
    "moneda":           '"$"#,##0',                       # $1,500
    "moneda_decimal":   '"$"#,##0.00',                    # $1,500.00
    "porcentaje":       '0%',                             # 15%
    "porcentaje_2dec":  '0.00%',                          # 15.00%
    "numero":           '#,##0',                          # 1,500
    "numero_decimal":   '#,##0.00',                       # 1,500.00
    "peso":             '#,##0.00 "Kgs"',                 # 2,540.00 Kgs
    "fecha":            'yyyy-mm-dd',
    "texto":            '@',
}


# ============================================================
# ANCHOS DE COLUMNA RECOMENDADOS
# ============================================================
ANCHOS_COLUMNA = {
    "estrecho": 10,
    "normal":   15,
    "medio":    20,
    "ancho":    30,
    "muy_ancho": 45,
}


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================
def aplicar_estilo_encabezado(celda, texto: str = None):
    """Aplica estilo de encabezado a una celda."""
    if texto is not None:
        celda.value = texto
    celda.font = FUENTES["encabezado"]
    celda.fill = RELLENOS["encabezado"]
    celda.alignment = ALINEACIONES["centro"]
    celda.border = BORDES["fino"]


def aplicar_estilo_celda(celda, formato: str = None, tipo_relleno: str = None):
    """Aplica estilo a una celda de datos."""
    celda.font = FUENTES["normal"]
    celda.border = BORDES["fino"]
    if formato:
        celda.number_format = FORMATOS.get(formato, formato)
    if tipo_relleno:
        celda.fill = RELLENOS.get(tipo_relleno, RELLENOS["fila_normal"])


def ajustar_ancho_columnas(ws, columnas_anchos: dict):
    """
    Ajusta el ancho de columnas.
    Args:
        ws: worksheet de openpyxl
        columnas_anchos: dict {1: 20, 2: 15, ...} (1-indexed)
    """
    for col_idx, ancho in columnas_anchos.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho


def congelar_paneles(ws, fila: int = 2, columna: int = 1):
    """Congela paneles en la posición indicada."""
    ws.freeze_panes = ws.cell(row=fila, column=columna)