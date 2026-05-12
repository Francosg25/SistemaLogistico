"""
Constructor de la hoja Sea del Excel de salida.
Maneja la regla CAPEX y la doble columna de %PCT (total y Summary).
"""
from openpyxl.worksheet.worksheet import Worksheet
from src.salida.estilos import (
    FUENTES, RELLENOS, ALINEACIONES, BORDES, FORMATOS,
    aplicar_estilo_encabezado, ajustar_ancho_columnas, congelar_paneles,
)


def construir_hoja_sea(ws: Worksheet, resultado_sea, costo_fijo: float = 2500) -> None:
    """
    Construye la hoja Sea.
    
    Estructura de columnas (críticas para que Summary haga XLOOKUP correctamente):
    A: BU | B: # Items | C: # Containers | D: Peso Total | E: Amount (USD)
    F: Total para SUM | G: Cost | H: %PCT (Summary - excluye Capex/MCS)
    """
    # ═══════════════════════════════════════════════════════
    # SECCIÓN 1: RESUMEN POR BU
    # ═══════════════════════════════════════════════════════
    headers = [
        "BU", "# Items", "# Contenedores", "Peso Total (Kgs)", 
        "Amount (USD)", "Total Cost", "Cost", "%PCT (Summary)"
    ]
    for col, header in enumerate(headers, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = header
        aplicar_estilo_encabezado(celda)
    
    df_resumen = resultado_sea.resumen_bu
    
    for idx, row in df_resumen.iterrows():
        fila = 2 + idx
        bu = row["BU"]
        es_excluido = bu in ("Capex", "MCS")
        
        ws.cell(row=fila, column=1, value=bu)
        ws.cell(row=fila, column=2, value=int(row.get("# Items", 0)))
        ws.cell(row=fila, column=3, value=int(row.get("# Contenedores", 0)))
        ws.cell(row=fila, column=4, value=float(row.get("Peso Total (Kgs)", 0)))
        ws.cell(row=fila, column=5, value=float(row.get("Amount (USD)", 0)))
        ws.cell(row=fila, column=6, value=float(row.get("Amount (USD)", 0)))
        ws.cell(row=fila, column=7, value=float(row.get("Amount (USD)", 0)))
        # Columna H: %PCT (Summary) - YA viene del Bloque 4 con Capex/MCS = 0
        ws.cell(row=fila, column=8, value=float(row.get("%PCT (Summary)", 0)))
        
        # Formatos
        ws.cell(row=fila, column=4).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=5).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=6).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=7).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=8).number_format = FORMATOS["porcentaje"]
        
        # Estilo especial para Capex/MCS
        relleno = RELLENOS["excluido"] if es_excluido else (
            RELLENOS["fila_alterna"] if idx % 2 == 0 else RELLENOS["fila_normal"]
        )
        for col in range(1, 9):
            celda = ws.cell(row=fila, column=col)
            celda.border = BORDES["fino"]
            celda.alignment = ALINEACIONES["centro"]
            celda.fill = relleno
    
    # ═══════════════════════════════════════════════════════
    # SECCIÓN 2: DETALLE POR ITEM (con fórmulas)
    # ═══════════════════════════════════════════════════════
    fila_detalle = 2 + len(df_resumen) + 3
    
    ws.cell(row=fila_detalle - 1, column=1, value="📋 DETALLE POR ITEM").font = FUENTES["subtitulo"]
    ws.cell(row=fila_detalle - 1, column=1).fill = RELLENOS["subtitulo"]
    ws.merge_cells(start_row=fila_detalle - 1, start_column=1,
                   end_row=fila_detalle - 1, end_column=8)
    
    headers_det = [
        "BU", "Item Code", "Container Number", "Total Gross Weight",
        "Peso Total Cont.", "%Pond", "Cost", "Es CAPEX"
    ]
    for col, header in enumerate(headers_det, start=1):
        celda = ws.cell(row=fila_detalle, column=col)
        celda.value = header
        aplicar_estilo_encabezado(celda)
    
    df_detalle = resultado_sea.detalle
    for idx, row in df_detalle.iterrows():
        fila = fila_detalle + 1 + idx
        es_capex = bool(row.get("Es CAPEX", False))
        
        ws.cell(row=fila, column=1, value=row.get("BU", ""))
        ws.cell(row=fila, column=2, value=row.get("Item Code", ""))
        ws.cell(row=fila, column=3, value=row.get("Container Number", ""))
        ws.cell(row=fila, column=4, value=float(row.get("Total Gross Weight", 0)))
        
        if es_capex:
            # CAPEX: peso 0, %Pond = 100%, Cost = $2,500
            ws.cell(row=fila, column=5, value=0)
            ws.cell(row=fila, column=6, value=1.0)
            ws.cell(row=fila, column=7, value=costo_fijo)
        else:
            # Normal: Fórmulas Excel auditables
            ws.cell(row=fila, column=5, value=f'=SUMIFS(D:D,C:C,C{fila})')
            ws.cell(row=fila, column=6, value=f'=IFERROR(D{fila}/E{fila},0)')
            ws.cell(row=fila, column=7, value=f'=F{fila}*{costo_fijo}')
        
        ws.cell(row=fila, column=8, value="SÍ" if es_capex else "NO")
        
        # Formatos
        ws.cell(row=fila, column=4).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=5).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=6).number_format = FORMATOS["porcentaje_2dec"]
        ws.cell(row=fila, column=7).number_format = FORMATOS["moneda"]
        
        # Color especial para CAPEX
        relleno = RELLENOS["capex"] if es_capex else (
            RELLENOS["fila_alterna"] if idx % 2 == 0 else RELLENOS["fila_normal"]
        )
        for col in range(1, 9):
            celda = ws.cell(row=fila, column=col)
            celda.border = BORDES["fino"]
            celda.alignment = ALINEACIONES["centro"]
            celda.fill = relleno
    
    # ═══════════════════════════════════════════════════════
    # AJUSTES
    # ═══════════════════════════════════════════════════════
    anchos = {1: 10, 2: 18, 3: 18, 4: 14, 5: 14, 6: 10, 7: 12, 8: 10}
    ajustar_ancho_columnas(ws, anchos)
    congelar_paneles(ws, fila=2, columna=2)
