"""Constructor de la hoja de Validaciones (reporte del Bloque 8)."""
from openpyxl.worksheet.worksheet import Worksheet
from src.salida.estilos import (
    FUENTES, RELLENOS, ALINEACIONES, BORDES,
    aplicar_estilo_encabezado, ajustar_ancho_columnas,
)


def construir_hoja_validaciones(ws: Worksheet, reporte_validacion) -> None:
    """Construye una hoja con todas las validaciones ejecutadas."""
    if reporte_validacion is None:
        ws["A1"] = "No se ejecutaron validaciones."
        return
    
    # Título
    ws.merge_cells("A1:F1")
    ws["A1"] = "✅ REPORTE DE VALIDACIONES AUTOMÁTICAS"
    ws["A1"].font = FUENTES["titulo"]
    ws["A1"].fill = RELLENOS["titulo"]
    ws["A1"].alignment = ALINEACIONES["centro"]
    ws.row_dimensions[1].height = 28
    
    # Resumen ejecutivo
    resumen = reporte_validacion.resumen()
    ws["A3"] = "Estado Global:"
    ws["B3"] = f"{resumen['emoji_global']} {resumen['estado_global']}"
    ws["A4"] = "Total Validaciones:"
    ws["B4"] = resumen["total_validaciones"]
    ws["A5"] = "🟢 OK:"
    ws["B5"] = resumen["ok"]
    ws["A6"] = "🟡 Warnings:"
    ws["B6"] = resumen["warnings"]
    ws["A7"] = "🔴 Errores:"
    ws["B7"] = resumen["errores"]
    ws["A8"] = "¿Puede exportar?"
    ws["B8"] = "✅ SÍ" if resumen["puede_exportar"] else "❌ NO"
    
    for fila in range(3, 9):
        ws.cell(row=fila, column=1).font = FUENTES["negrita"]
        ws.cell(row=fila, column=1).fill = RELLENOS["fila_alterna"]
    
    # Tabla de hallazgos
    fila_inicio = 10
    headers = ["Severidad", "Operación", "Regla", "Mensaje", "Acción Sugerida"]
    for col, header in enumerate(headers, start=1):
        celda = ws.cell(row=fila_inicio, column=col)
        celda.value = header
        aplicar_estilo_encabezado(celda)
    
    for idx, hallazgo in enumerate(reporte_validacion.hallazgos):
        fila = fila_inicio + 1 + idx
        ws.cell(row=fila, column=1, value=f"{hallazgo.severidad.emoji} {hallazgo.severidad.value}")
        ws.cell(row=fila, column=2, value=hallazgo.operacion.upper())
        ws.cell(row=fila, column=3, value=hallazgo.regla)
        ws.cell(row=fila, column=4, value=hallazgo.mensaje)
        ws.cell(row=fila, column=5, value=hallazgo.accion_sugerida or "")
        
        # Color por severidad
        severidad = hallazgo.severidad.value
        if severidad == "OK":
            relleno = RELLENOS["ok"]
        elif severidad == "WARNING":
            relleno = RELLENOS["warning"]
        elif severidad in ("ERROR", "CRITICAL"):
            relleno = RELLENOS["error"]
        else:
            relleno = RELLENOS["fila_normal"]
        
        for col in range(1, 6):
            celda = ws.cell(row=fila, column=col)
            celda.border = BORDES["fino"]
            celda.alignment = ALINEACIONES["izquierda"]
            celda.fill = relleno
    
    ajustar_ancho_columnas(ws, {1: 14, 2: 12, 3: 28, 4: 60, 5: 50})