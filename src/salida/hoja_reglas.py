"""Constructor de la hoja REGLAS_PROCESO (documentación)."""
from openpyxl.worksheet.worksheet import Worksheet
from src.salida.estilos import FUENTES, RELLENOS, ALINEACIONES, BORDES, ajustar_ancho_columnas


REGLAS_TEXTO = [
    ("📚 REGLAS DEL PROCESO DE CONSOLIDACIÓN LOGÍSTICA", "titulo"),
    ("", None),
    ("⚠️ REGLA CRÍTICA: FLEXIBILIDAD DE BUSINESS UNITS", "subtitulo"),
    ("🔴 NO ENCASILLARSE CON LOS MISMOS BU", "negrita"),
    ("Los Business Units pueden CAMBIAR mes a mes:", None),
    ("  • Pueden aparecer NUEVOS BU no vistos antes", None),
    ("  • Pueden DESAPARECER BU que existían", None),
    ("  • Las DESCRIPCIONES pueden cambiar", None),
    ("✅ SIEMPRE: Leer los BU directamente de los datos fuente", None),
    ("❌ NUNCA: Asumir que los BU serán iguales al mes anterior", None),
    ("", None),
    ("📤 REGLAS OUTBOUND (Exportaciones)", "subtitulo"),
    ("Costo Fijo: $1,500 USD por Reference", None),
    ("Agrupación: Por Reference (Waybill Number)", None),
    ("Inferencia BU: Segundo BU del patrón (ej: M01/M45 → M45)", None),
    ("Fórmula %Proportion: =Peso/SUMIFS(Peso,Reference,MiReference)", None),
    ("Fórmula Calc_Exp: =%Proportion × $1,500", None),
    ("", None),
    ("🚢 REGLAS SEA (Importaciones Marítimas)", "subtitulo"),
    ("Costo Fijo: $2,500 USD por contenedor", None),
    ("Agrupación: Por Container Number", None),
    ("Fórmula %Pond: =Peso/SUMIFS(Peso,Container,MiContainer)", None),
    ("Fórmula Cost: =%Pond × $2,500", None),
    ("", None),
    ("🔴 REGLA ESPECIAL CAPEX (Sea):", "negrita"),
    ("Identificación: Item Code contiene 'CAPEX'", None),
    ("Origen: NO viene en el reporte - se agrega MANUALMENTE en la UI", None),
    ("BU Asignado: 'Capex' | Peso: 0 | %Pond: 100% | Cost: $2,500", None),
    ("", None),
    ("📊 REGLA DE EXCLUSIÓN (Summary):", "negrita"),
    ("Para %PCT del Summary: Excluir 'Capex' y 'MCS'", None),
    ("Los montos absolutos sí se reportan completos", None),
    ("", None),
    ("🚛 REGLAS LAND (Importaciones Terrestres)", "subtitulo"),
    ("Costo Fijo: $1,200 USD por Reference", None),
    ("Agrupación: Por Reference", None),
    ("BU: Se toma DIRECTAMENTE de la columna BU del reporte (no se infiere)", None),
    ("BUs especiales: Machine, Miscelaneus (incluidos en %PCT)", None),
]


def construir_hoja_reglas(ws: Worksheet) -> None:
    """Construye la hoja REGLAS_PROCESO con la documentación del sistema."""
    for idx, (texto, estilo) in enumerate(REGLAS_TEXTO, start=1):
        celda = ws.cell(row=idx, column=1, value=texto)
        
        if estilo == "titulo":
            celda.font = FUENTES["titulo"]
            celda.fill = RELLENOS["titulo"]
            celda.alignment = ALINEACIONES["centro"]
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=6)
            ws.row_dimensions[idx].height = 26
        elif estilo == "subtitulo":
            celda.font = FUENTES["subtitulo"]
            celda.fill = RELLENOS["subtitulo"]
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=6)
        elif estilo == "negrita":
            celda.font = FUENTES["negrita"]
        else:
            celda.font = FUENTES["normal"]
    
    ajustar_ancho_columnas(ws, {1: 80})