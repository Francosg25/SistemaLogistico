"""Constructor de la hoja Land del Excel de salida."""
from openpyxl.worksheet.worksheet import Worksheet
from src.salida.estilos import (
    FUENTES, RELLENOS, ALINEACIONES, BORDES, FORMATOS,
    aplicar_estilo_encabezado, ajustar_ancho_columnas, congelar_paneles,
)


def construir_hoja_land(ws: Worksheet, resultado_land, costo_fijo: float = 1200) -> None:
    """
    Estructura de columnas para que Summary haga XLOOKUP:
    A: BU | B: # References | C: # Items | D: Peso Total | E: Monto Total (USD) 
    F: Total para SUM | G: %PCT
    """
    # ═══════════════════════════════════════════════════════
    # SECCIÓN 1: RESUMEN POR BU
    # ═══════════════════════════════════════════════════════
    headers = [
        "BU", "# References", "# Items", "Peso Total (Kgs)", 
        "Monto Total (USD)", "Total Cost", "%PCT"
    ]
    for col, header in enumerate(headers, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = header
        aplicar_estilo_encabezado(celda)
    
    df_resumen = resultado_land.resumen_bu
    for idx, row in df_resumen.iterrows():
        fila = 2 + idx
        bu = row["BU"]
        es_especial = bu in ("Machine", "Miscelaneus")
        
        ws.cell(row=fila, column=1, value=bu)
        ws.cell(row=fila, column=2, value=int(row.get("# References", 0)))
        ws.cell(row=fila, column=3, value=int(row.get("# Items", 0)))
        ws.cell(row=fila, column=4, value=float(row.get("Peso Total (Kgs)", 0)))
        ws.cell(row=fila, column=5, value=float(row.get("Monto Total (USD)", 0)))
        ws.cell(row=fila, column=6, value=float(row.get("Monto Total (USD)", 0)))
        ws.cell(row=fila, column=7, value=float(row.get("%PCT", 0)))
        
        ws.cell(row=fila, column=4).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=5).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=6).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=7).number_format = FORMATOS["porcentaje"]
        
        relleno = RELLENOS["capex"] if es_especial else (
            RELLENOS["fila_alterna"] if idx % 2 == 0 else RELLENOS["fila_normal"]
        )
        for col in range(1, 8):
            celda = ws.cell(row=fila, column=col)
            celda.border = BORDES["fino"]
            celda.alignment = ALINEACIONES["centro"]
            celda.fill = relleno
    
    # ═══════════════════════════════════════════════════════
    # SECCIÓN 2: DETALLE POR ITEM (con fórmulas)
    # ═══════════════════════════════════════════════════════
    fila_detalle = 2 + len(df_resumen) + 3
    
    ws.cell(row=fila_detalle - 1, column=1, value="📋 DETALLE POR ITEM").font = FUENTES["subtitulo"]
    ws.cell(row=fila_detalle - 1, column=1).fill = RELLENOS["subtitulo"]
    ws.merge_cells(start_row=fila_detalle - 1, start_column=1,
                   end_row=fila_detalle - 1, end_column=8)
    
    headers_det = ["Reference", "BU", "No. Parte", "Peso Bruto (Kgs)",
                   "Peso Total Ref", "%Pond", "Fix Cost", "Cost"]
    for col, header in enumerate(headers_det, start=1):
        celda = ws.cell(row=fila_detalle, column=col)
        celda.value = header
        aplicar_estilo_encabezado(celda)
    
    df_detalle = resultado_land.detalle
    for idx, row in df_detalle.iterrows():
        fila = fila_detalle + 1 + idx
        
        ws.cell(row=fila, column=1, value=row.get("Reference", ""))
        ws.cell(row=fila, column=2, value=row.get("BU", ""))
        ws.cell(row=fila, column=3, value=row.get("No. Parte Prov.", ""))
        ws.cell(row=fila, column=4, value=float(row.get("Peso Bruto (Kgs)", 0)))
        # Fórmulas
        ws.cell(row=fila, column=5, value=f'=SUMIFS(D:D,A:A,A{fila})')
        ws.cell(row=fila, column=6, value=f'=IFERROR(D{fila}/E{fila},0)')
        ws.cell(row=fila, column=7, value=costo_fijo)
        ws.cell(row=fila, column=8, value=f'=F{fila}*G{fila}')
        
        ws.cell(row=fila, column=4).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=5).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=6).number_format = FORMATOS["porcentaje_2dec"]
        ws.cell(row=fila, column=7).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=8).number_format = FORMATOS["moneda"]
        
        relleno = RELLENOS["fila_alterna"] if idx % 2 == 0 else RELLENOS["fila_normal"]
        for col in range(1, 9):
            celda = ws.cell(row=fila, column=col)
            celda.border = BORDES["fino"]
            celda.alignment = ALINEACIONES["centro"]
            celda.fill = relleno
    
    anchos = {1: 20, 2: 14, 3: 20, 4: 16, 5: 16, 6: 12, 7: 12, 8: 14}
    ajustar_ancho_columnas(ws, anchos)
    congelar_paneles(ws, fila=2, columna=2)