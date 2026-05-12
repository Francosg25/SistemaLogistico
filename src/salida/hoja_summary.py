"""
Constructor de la hoja Summary del Excel de salida.

CARACTERÍSTICAS CRÍTICAS:
- Usa fórmulas Excel reales (XLOOKUP, SUMIFS) - NO valores hard-coded
- El usuario puede auditar los cálculos directamente
- Replica la estructura exacta de tu hoja Summary actual (B3:H12)
"""
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import List

from src.salida.estilos import (
    FUENTES, RELLENOS, ALINEACIONES, BORDES, FORMATOS,
    aplicar_estilo_encabezado, aplicar_estilo_celda,
    ajustar_ancho_columnas, congelar_paneles,
)


def construir_hoja_summary(ws: Worksheet, bus_orden: List[str]) -> None:
    """
    Construye la hoja Summary con fórmulas Excel que referencian las hojas
    Outbound, Sea y Land.
    
    Estructura final (replicando Summary!B3:H12 del archivo original):
        B3:H6   → Tabla de %PCT  (encabezado + 3 filas: Sea/Land/Outbound)
        B9:H12  → Tabla de Montos (encabezado + 3 filas: Sea/Land/Outbound)
    
    Args:
        ws: Worksheet de openpyxl donde construir
        bus_orden: Lista ordenada de BUs detectados (ej: ['M01', 'M19', 'M23', 'M45', 'M46'])
    """
    # ─────────────────────────────────────────────────────────
    # CONFIGURACIÓN INICIAL
    # ─────────────────────────────────────────────────────────
    # Filtrar BUs que SÍ van al %PCT (excluir Capex y MCS)
    bus_pct = [bu for bu in bus_orden if bu not in ("Capex", "MCS")]
    
    # ─────────────────────────────────────────────────────────
    # TÍTULO PRINCIPAL
    # ─────────────────────────────────────────────────────────
    ws.merge_cells("B1:H1")
    ws["B1"] = "📊 SUMMARY CONSOLIDADO - Consolidación Logística"
    ws["B1"].font = FUENTES["titulo"]
    ws["B1"].fill = RELLENOS["titulo"]
    ws["B1"].alignment = ALINEACIONES["centro"]
    ws.row_dimensions[1].height = 28
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN 1: TABLA DE %PCT (filas 3-6)
    # ─────────────────────────────────────────────────────────
    # Fila 3: Encabezados (Type + BUs)
    ws["C3"] = "Type"
    aplicar_estilo_encabezado(ws["C3"])
    
    for idx, bu in enumerate(bus_pct):
        col_letra = get_column_letter(4 + idx)  # D=4, E=5, F=6...
        celda = ws[f"{col_letra}3"]
        celda.value = bu
        aplicar_estilo_encabezado(celda)
    
    # Fila 4: Sea %PCT (con fórmula XLOOKUP referenciando hoja Sea)
    ws["C4"] = "Sea %PCT"
    ws["C4"].font = FUENTES["negrita"]
    ws["C4"].fill = RELLENOS["fila_alterna"]
    ws["C4"].border = BORDES["fino"]
    
    for idx, bu in enumerate(bus_pct):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}4"]
        # Fórmula XLOOKUP que busca el BU en la hoja Sea y devuelve el %PCT
        celda.value = (
            f'=IFERROR(XLOOKUP("{bu}",Sea!A:A,Sea!H:H,0),0)'
        )
        celda.number_format = FORMATOS["porcentaje"]
        celda.border = BORDES["fino"]
        celda.alignment = ALINEACIONES["centro"]
        celda.fill = RELLENOS["fila_alterna"]
    
    # Fila 5: Land %PCT
    ws["C5"] = "Land %PCT"
    ws["C5"].font = FUENTES["negrita"]
    ws["C5"].border = BORDES["fino"]
    
    for idx, bu in enumerate(bus_pct):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}5"]
        celda.value = (
            f'=IFERROR(XLOOKUP("{bu}",Land!A:A,Land!G:G,0),0)'
        )
        celda.number_format = FORMATOS["porcentaje"]
        celda.border = BORDES["fino"]
        celda.alignment = ALINEACIONES["centro"]
    
    # Fila 6: Outbound %PCT
    ws["C6"] = "Outbound %PCT"
    ws["C6"].font = FUENTES["negrita"]
    ws["C6"].fill = RELLENOS["fila_alterna"]
    ws["C6"].border = BORDES["fino"]
    
    for idx, bu in enumerate(bus_pct):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}6"]
        celda.value = (
            f'=IFERROR(XLOOKUP("{bu}",Outbound!A:A,Outbound!G:G,0),0)'
        )
        celda.number_format = FORMATOS["porcentaje"]
        celda.border = BORDES["fino"]
        celda.alignment = ALINEACIONES["centro"]
        celda.fill = RELLENOS["fila_alterna"]
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN 2: TABLA DE MONTOS $ (filas 9-12)
    # ─────────────────────────────────────────────────────────
    # Fila 9: Encabezados (Arg. Var $ + Total + BUs completos incluyendo Capex/MCS)
    ws["B9"] = "Viewer"
    ws["C9"] = "Arg. Var $"
    ws["B9"].font = FUENTES["encabezado"]
    ws["B9"].fill = RELLENOS["encabezado"]
    ws["B9"].alignment = ALINEACIONES["centro"]
    ws["B9"].border = BORDES["fino"]
    aplicar_estilo_encabezado(ws["C9"])
    
    # Para la tabla de montos SÍ incluimos todos los BUs (incluyendo Capex/MCS)
    for idx, bu in enumerate(bus_orden):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}9"]
        celda.value = bu
        aplicar_estilo_encabezado(celda)
        # Si es Capex o MCS, color diferente
        if bu in ("Capex", "MCS"):
            celda.fill = RELLENOS["excluido"]
            celda.font = FUENTES["negrita"]
    
    # ── Fila 10: Sea con fórmulas SUMIFS
    ws["B10"] = "Sea"
    ws["B10"].font = FUENTES["negrita"]
    ws["B10"].fill = RELLENOS["fila_alterna"]
    ws["B10"].border = BORDES["fino"]
    ws["B10"].alignment = ALINEACIONES["centro"]
    
    # Columna C10: Total Sea (suma de toda la columna Cost de la hoja Sea)
    ws["C10"] = '=SUM(Sea!G:G)'
    ws["C10"].number_format = FORMATOS["moneda"]
    ws["C10"].font = FUENTES["negrita"]
    ws["C10"].border = BORDES["fino"]
    ws["C10"].fill = RELLENOS["fila_alterna"]
    ws["C10"].alignment = ALINEACIONES["centro"]
    
    # Columnas D-? : Montos por BU (con XLOOKUP a la hoja Sea)
    for idx, bu in enumerate(bus_orden):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}10"]
        celda.value = f'=IFERROR(XLOOKUP("{bu}",Sea!A:A,Sea!E:E,0),0)'
        celda.number_format = FORMATOS["moneda"]
        celda.border = BORDES["fino"]
        celda.alignment = ALINEACIONES["centro"]
        celda.fill = RELLENOS["fila_alterna"]
    
    # ── Fila 11: Land
    ws["B11"] = "Land"
    ws["B11"].font = FUENTES["negrita"]
    ws["B11"].border = BORDES["fino"]
    ws["B11"].alignment = ALINEACIONES["centro"]
    
    ws["C11"] = '=SUM(Land!F:F)'
    ws["C11"].number_format = FORMATOS["moneda"]
    ws["C11"].font = FUENTES["negrita"]
    ws["C11"].border = BORDES["fino"]
    ws["C11"].alignment = ALINEACIONES["centro"]
    
    for idx, bu in enumerate(bus_orden):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}11"]
        celda.value = f'=IFERROR(XLOOKUP("{bu}",Land!A:A,Land!E:E,0),0)'
        celda.number_format = FORMATOS["moneda"]
        celda.border = BORDES["fino"]
        celda.alignment = ALINEACIONES["centro"]
    
    # ── Fila 12: Outbound
    ws["B12"] = "Outbound"
    ws["B12"].font = FUENTES["negrita"]
    ws["B12"].fill = RELLENOS["fila_alterna"]
    ws["B12"].border = BORDES["fino"]
    ws["B12"].alignment = ALINEACIONES["centro"]
    
    ws["C12"] = '=SUM(Outbound!F:F)'
    ws["C12"].number_format = FORMATOS["moneda"]
    ws["C12"].font = FUENTES["negrita"]
    ws["C12"].border = BORDES["fino"]
    ws["C12"].fill = RELLENOS["fila_alterna"]
    ws["C12"].alignment = ALINEACIONES["centro"]
    
    for idx, bu in enumerate(bus_orden):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}12"]
        celda.value = f'=IFERROR(XLOOKUP("{bu}",Outbound!A:A,Outbound!E:E,0),0)'
        celda.number_format = FORMATOS["moneda"]
        celda.border = BORDES["fino"]
        celda.alignment = ALINEACIONES["centro"]
        celda.fill = RELLENOS["fila_alterna"]
    
    # ─────────────────────────────────────────────────────────
    # FILA DE TOTAL GENERAL (fila 14)
    # ─────────────────────────────────────────────────────────
    ws["B14"] = "GRAN TOTAL"
    ws["B14"].font = FUENTES["total"]
    ws["B14"].fill = RELLENOS["total"]
    ws["B14"].alignment = ALINEACIONES["centro"]
    ws["B14"].border = BORDES["medio"]
    
    ws["C14"] = "=SUM(C10:C12)"
    ws["C14"].number_format = FORMATOS["moneda"]
    ws["C14"].font = FUENTES["total"]
    ws["C14"].fill = RELLENOS["total"]
    ws["C14"].alignment = ALINEACIONES["centro"]
    ws["C14"].border = BORDES["medio"]
    
    for idx, bu in enumerate(bus_orden):
        col_letra = get_column_letter(4 + idx)
        celda = ws[f"{col_letra}14"]
        celda.value = f"=SUM({col_letra}10:{col_letra}12)"
        celda.number_format = FORMATOS["moneda"]
        celda.font = FUENTES["total"]
        celda.fill = RELLENOS["total"]
        celda.alignment = ALINEACIONES["centro"]
        celda.border = BORDES["medio"]
    
    # ─────────────────────────────────────────────────────────
    # CAPTION / NOTAS AL PIE
    # ─────────────────────────────────────────────────────────
    ws["B16"] = (
        "ℹ️ Notas: %PCT Sea excluye Capex y MCS (regla del negocio). "
        "Los montos en la sección inferior SÍ incluyen todos los BUs. "
        "Todas las fórmulas hacen referencia a las hojas Outbound, Sea y Land."
    )
    ws["B16"].font = FUENTES["caption"]
    ws.merge_cells("B16:H16")
    
    # ─────────────────────────────────────────────────────────
    # AJUSTES VISUALES FINALES
    # ─────────────────────────────────────────────────────────
    anchos = {1: 3, 2: 15, 3: 18}  # A=3, B=15, C=18
    for idx in range(len(bus_orden)):
        anchos[4 + idx] = 14  # BUs
    ajustar_ancho_columnas(ws, anchos)
    
    # Altura especial para filas de encabezado
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[14].height = 25
