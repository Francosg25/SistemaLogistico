"""
Constructor de la hoja Outbound del Excel de salida.

Estructura:
- Columna A: BU (clave para XLOOKUP desde Summary)
- Columna B-F: Métricas resumen del BU
- Columna G: %PCT (referenciado por Summary)
- A partir de fila N: detalle completo con fórmulas
"""
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from src.salida.estilos import (
    FUENTES, RELLENOS, ALINEACIONES, BORDES, FORMATOS,
    aplicar_estilo_encabezado, ajustar_ancho_columnas, congelar_paneles,
)


def construir_hoja_outbound(ws: Worksheet, resultado_outbound, costo_fijo: float = 1500) -> None:
    """
    Construye la hoja Outbound con dos secciones:
    1. Resumen por BU (filas 1-N) - referenciado por Summary
    2. Detalle completo con fórmulas (a partir de fila N+3)
    """
    # ═══════════════════════════════════════════════════════
    # SECCIÓN 1: RESUMEN POR BU (filas 1-N)
    # ═══════════════════════════════════════════════════════
    # Encabezados
    headers_resumen = ["BU", "# References", "# Items", "Peso Total", "Log. Exp (USD)", "Calc Total", "%PCT"]
    for col, header in enumerate(headers_resumen, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = header
        aplicar_estilo_encabezado(celda)
    
    # Datos del resumen
    df_resumen = resultado_outbound.resumen_bu
    fila_inicio_resumen = 2
    
    for idx, row in df_resumen.iterrows():
        fila = fila_inicio_resumen + idx
        ws.cell(row=fila, column=1, value=row["BU"])
        ws.cell(row=fila, column=2, value=row.get("# References", 0))
        ws.cell(row=fila, column=3, value=row.get("# Items", 0))
        ws.cell(row=fila, column=4, value=float(row.get("Peso Total (Kgs)", 0)))
        # Columna E: Log. Exp (referenciada por Summary!Outbound!E:E)
        ws.cell(row=fila, column=5, value=float(row.get("Log. Exp (USD)", 0)))
        # Columna F: Total para SUM(Outbound!F:F) en Summary
        ws.cell(row=fila, column=6, value=float(row.get("Log. Exp (USD)", 0)))
        # Columna G: %PCT (referenciada por Summary!Outbound!G:G)
        ws.cell(row=fila, column=7, value=float(row.get("%PCT", 0)))
        
        # Formatos
        ws.cell(row=fila, column=4).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=5).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=6).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=7).number_format = FORMATOS["porcentaje"]
        
        for col in range(1, 8):
            ws.cell(row=fila, column=col).border = BORDES["fino"]
            ws.cell(row=fila, column=col).alignment = ALINEACIONES["centro"]
    
    # ═══════════════════════════════════════════════════════
    # SECCIÓN 2: DETALLE COMPLETO (con fórmulas)
    # ═══════════════════════════════════════════════════════
    fila_detalle_inicio = fila_inicio_resumen + len(df_resumen) + 3
    
    # Título de la sección
    ws.cell(row=fila_detalle_inicio - 1, column=1, value="📋 DETALLE COMPLETO POR ITEM").font = FUENTES["subtitulo"]
    ws.cell(row=fila_detalle_inicio - 1, column=1).fill = RELLENOS["subtitulo"]
    ws.merge_cells(start_row=fila_detalle_inicio - 1, start_column=1,
                   end_row=fila_detalle_inicio - 1, end_column=10)
    
    # Encabezados del detalle
    headers_detalle = [
        "Reference", "BU (Inferido)", "Item", "Qty Pzas", 
        "Gross Weight", "Peso Total Ref", "%Proportion", "Fix Cost", "Calc_Exp"
    ]
    for col, header in enumerate(headers_detalle, start=1):
        celda = ws.cell(row=fila_detalle_inicio, column=col)
        celda.value = header
        aplicar_estilo_encabezado(celda)
    
    # Datos del detalle con FÓRMULAS Excel reales
    df_detalle = resultado_outbound.detalle
    for idx, row in df_detalle.iterrows():
        fila = fila_detalle_inicio + 1 + idx
        
        ws.cell(row=fila, column=1, value=row.get("Reference", ""))
        ws.cell(row=fila, column=2, value=row.get("BU (Inferido)", ""))
        ws.cell(row=fila, column=3, value=row.get("Item", ""))
        ws.cell(row=fila, column=4, value=float(row.get("Qty Pzas", 0)) if row.get("Qty Pzas") else 0)
        ws.cell(row=fila, column=5, value=float(row.get("Gross Weight", 0)))
        
        # Columna F: Peso Total Reference (FÓRMULA SUMIFS - auditable)
        ws.cell(row=fila, column=6, value=(
            f'=SUMIFS(E:E,A:A,A{fila})'
        ))
        
        # Columna G: %Proportion (FÓRMULA)
        ws.cell(row=fila, column=7, value=(
            f'=IFERROR(E{fila}/F{fila},0)'
        ))
        
        # Columna H: Fix Cost
        ws.cell(row=fila, column=8, value=costo_fijo)
        
        # Columna I: Calc_Exp (FÓRMULA)
        ws.cell(row=fila, column=9, value=(
            f'=G{fila}*H{fila}'
        ))
        
        # Formatos
        ws.cell(row=fila, column=4).number_format = FORMATOS["numero"]
        ws.cell(row=fila, column=5).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=6).number_format = FORMATOS["peso"]
        ws.cell(row=fila, column=7).number_format = FORMATOS["porcentaje_2dec"]
        ws.cell(row=fila, column=8).number_format = FORMATOS["moneda"]
        ws.cell(row=fila, column=9).number_format = FORMATOS["moneda"]
        
        # Bordes y alineación
        for col in range(1, 10):
            celda = ws.cell(row=fila, column=col)
            celda.border = BORDES["fino"]
            celda.alignment = ALINEACIONES["centro"]
            # Fila alterna
            if idx % 2 == 0:
                celda.fill = RELLENOS["fila_alterna"]
    
    # ═══════════════════════════════════════════════════════
    # AJUSTES VISUALES
    # ═══════════════════════════════════════════════════════
    anchos = {1: 22, 2: 12, 3: 20, 4: 10, 5: 14, 6: 16, 7: 13, 8: 12, 9: 14}
    ajustar_ancho_columnas(ws, anchos)
    congelar_paneles(ws, fila=2, columna=2)