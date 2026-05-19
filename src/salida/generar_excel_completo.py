"""
═══════════════════════════════════════════════════════════════
GENERADOR DE EXCEL CON FÓRMULAS VIVAS — Réplica del original
═══════════════════════════════════════════════════════════════
Produce un archivo XLSX con:
  • Hoja 'Summary'  → XLOOKUP IDÉNTICO al Excel original
  • Hoja 'Outbound' → fórmulas BP/BQ + tabla BC:BE + pivote BS:BU
  • Hoja 'Sea'      → fórmulas CR/CS + pivote CU:CY (excluye Capex/MCS)
  • Hoja 'Land'     → fórmulas BY/BZ + pivote CC:CE (excluye Misc/Machine)

🎯 USUARIO PUEDE:
  • Abrir el archivo y ver TODAS las fórmulas (no valores pegados)
  • Modificar tarifas y todo se recalcula automáticamente
  • Auditar paso a paso igual que en su Excel actual
═══════════════════════════════════════════════════════════════
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional, List, Dict, Any

# ════════════════════════════════════════════════════════════
# ESTILOS
# ════════════════════════════════════════════════════════════
FONT_HDR = Font(bold=True, color="FFFFFF", size=11)
FILL_HDR = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
FILL_FORMULA = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
FILL_TARIFA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FONT_BOLD = Font(bold=True)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
FMT_MONTO = '"$"#,##0.00'
FMT_PCT = "0.00%"
FMT_PESO = "#,##0.000"

# ════════════════════════════════════════════════════════════
# Constantes del layout (replican el Excel original)
# ════════════════════════════════════════════════════════════
BUS_SUMMARY_DEFAULT = ["M01", "M19", "M23", "M45", "M46"]


# ════════════════════════════════════════════════════════════
# HOJA OUTBOUND
# ════════════════════════════════════════════════════════════
def _hoja_outbound(
    wb: Workbook,
    df: pd.DataFrame,
    df_tarifas: Optional[pd.DataFrame],
    tarifa_default: float,
    bus_summary: List[str],
) -> Dict[str, str]:
    """
    Layout (replica el Excel original):
      • Cols B-J  = datos del item
      • Cols L-M  = TABLA DE TARIFAS (Reference, Fix Cost)
      • Cols O-T  = encabezados de cálculo
      • Col  P    = %Proportion (fórmula)
      • Col  Q    = Calc_Exp    (fórmula con XLOOKUP)
      • Col  S-U  = RESUMEN POR BU (BU, Log. Exp, %PCT)
    
    Returns: dict con rangos para que Summary los referencie.
    """
    ws = wb.create_sheet("Outbound")

    # ───── 1. TABLA DE TARIFAS (cols L:M) ─────
    ws["L1"] = "📋 Tabla de Tarifas"
    ws["L1"].font = FONT_HDR
    ws["L1"].fill = FILL_HDR
    ws.merge_cells("L1:M1")

    ws["L2"] = "Reference"
    ws["M2"] = "Fix Cost"
    for c in (ws["L2"], ws["M2"]):
        c.font = FONT_HDR
        c.fill = FILL_HDR

    if df_tarifas is not None and len(df_tarifas) > 0:
        for i, row in enumerate(df_tarifas.itertuples(index=False), start=3):
            ws.cell(row=i, column=12, value=str(row[0]))
            ws.cell(row=i, column=13, value=float(row[1]))
            ws.cell(row=i, column=13).number_format = FMT_MONTO
        fin_tarifas = 2 + len(df_tarifas)
        usar_xlookup = True
    else:
        # No hay tabla → escribir solo el default como "tabla virtual"
        ws["L3"] = "DEFAULT"
        ws["M3"] = tarifa_default
        ws["M3"].number_format = FMT_MONTO
        ws["M3"].fill = FILL_TARIFA
        ws["M3"].font = FONT_BOLD
        fin_tarifas = 3
        usar_xlookup = False

    rango_ref = f"$L$3:$L${fin_tarifas}"
    rango_cost = f"$M$3:$M${fin_tarifas}"

    # ───── 2. ENCABEZADOS DE DATOS (fila 5) ─────
    headers = [
        ("B5", "Inbound/Outbound"),
        ("C5", "Method"),
        ("D5", "Reference"),
        ("E5", "Customer"),
        ("F5", "BU"),
        ("G5", "Item"),
        ("H5", "Qty Pzas"),
        ("I5", "Gross Weight"),
        ("J5", "%Proportion"),
        ("K5", "Calc_Exp"),
    ]
    for cell_addr, val in headers:
        c = ws[cell_addr]
        c.value = val
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    # ───── 3. MAPEAR COLUMNAS Y ESCRIBIR DATOS ─────
    col_waybill = next((c for c in ["Reference", "Waybill Number", "Waybill"] if c in df.columns), "Reference")
    col_cust = next((c for c in ["Customer", "Cliente"] if c in df.columns), None)
    col_bu = "BU Final" if "BU Final" in df.columns else "BU"
    col_item = "Item"
    col_cant = next((c for c in ["Cantidad", "Pieces", "Qty Pzas"] if c in df.columns), None)
    col_peso = "Peso Bruto"

    fila_ini = 6
    n = len(df)
    fila_fin = fila_ini + n - 1

    for i, (_, row) in enumerate(df.iterrows()):
        r = fila_ini + i
        ws.cell(row=r, column=2, value="Outbound")
        ws.cell(row=r, column=3, value="Land")
        ws.cell(row=r, column=4, value=str(row[col_waybill]))
        ws.cell(row=r, column=5, value=str(row[col_cust]) if col_cust and pd.notna(row.get(col_cust)) else "")
        ws.cell(row=r, column=6, value=str(row[col_bu]) if pd.notna(row.get(col_bu)) else "")
        ws.cell(row=r, column=7, value=str(row[col_item]))
        ws.cell(row=r, column=8, value=float(row[col_cant]) if col_cant and pd.notna(row.get(col_cant)) else 0)
        ws.cell(row=r, column=9, value=float(row[col_peso]) if pd.notna(row[col_peso]) else 0)
        ws.cell(row=r, column=9).number_format = FMT_PESO

        # 🔑 J = %Proportion (réplica BP9 del Excel original)
        ws.cell(row=r, column=10).value = (
            f"=I{r}/SUMIFS($I${fila_ini}:$I${fila_fin},$D${fila_ini}:$D${fila_fin},D{r})"
        )
        ws.cell(row=r, column=10).number_format = FMT_PCT
        ws.cell(row=r, column=10).fill = FILL_FORMULA

        # 🔑 K = Calc_Exp (réplica BQ9: XLOOKUP * %Proportion)
        if usar_xlookup:
            ws.cell(row=r, column=11).value = (
                f"=IFERROR(XLOOKUP(D{r},{rango_ref},{rango_cost},{tarifa_default}),{tarifa_default})*J{r}"
            )
        else:
            ws.cell(row=r, column=11).value = f"=$M$3*J{r}"
        ws.cell(row=r, column=11).number_format = FMT_MONTO
        ws.cell(row=r, column=11).fill = FILL_FORMULA

    # ───── 4. RESUMEN POR BU (cols N:P) — réplica BS:BU del Excel ─────
    ws["N5"] = "📊 Resumen por BU"
    ws["N5"].font = FONT_HDR
    ws["N5"].fill = FILL_HDR
    ws.merge_cells("N5:P5")

    for col_letter, val in zip(["N", "O", "P"], ["BU", "Log. Exp", "%PCT"]):
        c = ws[f"{col_letter}6"]
        c.value = val
        c.font = FONT_HDR
        c.fill = FILL_HDR

    bu_ini = 7
    for i, bu in enumerate(bus_summary):
        r = bu_ini + i
        ws.cell(row=r, column=14, value=bu)  # N
        ws.cell(row=r, column=15).value = (
            f"=SUMIFS($K${fila_ini}:$K${fila_fin},$F${fila_ini}:$F${fila_fin},N{r})"
        )
        ws.cell(row=r, column=15).number_format = FMT_MONTO

    bu_fin = bu_ini + len(bus_summary) - 1
    # Fila Total
    r_total = bu_fin + 1
    ws.cell(row=r_total, column=14, value="Total").font = FONT_BOLD
    ws.cell(row=r_total, column=15).value = f"=SUM(O{bu_ini}:O{bu_fin})"
    ws.cell(row=r_total, column=15).number_format = FMT_MONTO
    ws.cell(row=r_total, column=15).font = FONT_BOLD
    ws.cell(row=r_total, column=15).fill = FILL_TOTAL

    # %PCT (col P)
    for i, bu in enumerate(bus_summary):
        r = bu_ini + i
        ws.cell(row=r, column=16).value = f"=O{r}/$O${r_total}"
        ws.cell(row=r, column=16).number_format = FMT_PCT
    ws.cell(row=r_total, column=16, value=1.0).number_format = FMT_PCT
    ws.cell(row=r_total, column=16).font = FONT_BOLD
    ws.cell(row=r_total, column=16).fill = FILL_TOTAL

    # ───── 5. Anchos ─────
    widths = {1: 3, 2: 12, 3: 9, 4: 24, 5: 28, 6: 8, 7: 22, 8: 10, 9: 14, 10: 13, 11: 13,
              12: 24, 13: 13, 14: 14, 15: 14, 16: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    return {
        "bu_range": f"$N${bu_ini}:$N${bu_fin}",
        "amount_range": f"$O${bu_ini}:$O${bu_fin}",
        "pct_range": f"$P${bu_ini}:$P${bu_fin}",
    }


# ════════════════════════════════════════════════════════════
# HOJA SEA
# ════════════════════════════════════════════════════════════
def _hoja_sea(
    wb: Workbook,
    df: pd.DataFrame,
    tarifa: float,
    bus_summary: List[str],
    bus_excluir_pct: List[str] = None,
) -> Dict[str, str]:
    """
    Layout (réplica del Excel original):
      • Cell I2 = tarifa fija (réplica $CS$2)
      • Cols B:G = datos + %Pond + Cost
      • Cols I:K = Resumen por BU con %PCT (excluye Capex/MCS)
    """
    if bus_excluir_pct is None:
        bus_excluir_pct = ["Capex", "MCS"]

    ws = wb.create_sheet("Sea")

    # ───── Tarifa fija destacada ─────
    ws["H2"] = "Tarifa Container:"
    ws["H2"].font = FONT_BOLD
    ws["I2"] = tarifa
    ws["I2"].number_format = FMT_MONTO
    ws["I2"].fill = FILL_TARIFA
    ws["I2"].font = Font(bold=True, color="C00000")

    # ───── Encabezados (fila 5) ─────
    headers = [("B5", "BU"), ("C5", "Item Code"), ("D5", "Container Number"),
               ("E5", "Total Gross Weight"), ("F5", "%Pond"), ("G5", "Cost")]
    for addr, val in headers:
        c = ws[addr]
        c.value = val; c.font = FONT_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")

    # ───── Datos ─────
    col_bu = "BU Final" if "BU Final" in df.columns else "BU"
    col_item = "Item"
    col_cont = "Container"
    col_peso = "Peso Bruto"

    fila_ini, n = 6, len(df)
    fila_fin = fila_ini + n - 1

    for i, (_, row) in enumerate(df.iterrows()):
        r = fila_ini + i
        ws.cell(row=r, column=2, value=str(row.get(col_bu, "")))
        ws.cell(row=r, column=3, value=str(row[col_item]))
        ws.cell(row=r, column=4, value=str(row[col_cont]))
        ws.cell(row=r, column=5, value=float(row[col_peso]) if pd.notna(row[col_peso]) else 0)
        ws.cell(row=r, column=5).number_format = FMT_PESO

        # 🔑 F = %Pond (réplica CR6)
        ws.cell(row=r, column=6).value = (
            f"=IFERROR(E{r}/SUMIFS($E${fila_ini}:$E${fila_fin},$D${fila_ini}:$D${fila_fin},D{r}),100%)"
        )
        ws.cell(row=r, column=6).number_format = FMT_PCT
        ws.cell(row=r, column=6).fill = FILL_FORMULA

        # 🔑 G = Cost (réplica CS6 = CR6 * $CS$2)
        ws.cell(row=r, column=7).value = f"=F{r}*$I$2"
        ws.cell(row=r, column=7).number_format = FMT_MONTO
        ws.cell(row=r, column=7).fill = FILL_FORMULA

    # ───── Resumen por BU (cols I:K) ─────
    ws["I5"] = "📊 Resumen por BU"
    ws["I5"].font = FONT_HDR; ws["I5"].fill = FILL_HDR
    ws.merge_cells("I5:K5")

    for col_letter, val in zip(["I", "J", "K"], ["BU", "Amount (USD)", "%PCT"]):
        c = ws[f"{col_letter}6"]; c.value = val; c.font = FONT_HDR; c.fill = FILL_HDR

    # Ordenar BUs: válidos primero, excluidos al final
    bus_en_df = sorted(df[col_bu].dropna().unique().tolist())
    bus_validos = [b for b in bus_en_df if b not in bus_excluir_pct]
    bus_excl_present = [b for b in bus_en_df if b in bus_excluir_pct]
    bus_orden = bus_validos + bus_excl_present

    bu_ini = 7
    fila_validos_fin = bu_ini + len(bus_validos) - 1

    for i, bu in enumerate(bus_orden):
        r = bu_ini + i
        ws.cell(row=r, column=9, value=bu)
        # SUMIFS por BU
        ws.cell(row=r, column=10).value = (
            f"=SUMIFS($G${fila_ini}:$G${fila_fin},$B${fila_ini}:$B${fila_fin},I{r})"
        )
        ws.cell(row=r, column=10).number_format = FMT_MONTO

        # 🔑 %PCT EXCLUYE Capex/MCS (réplica CY7 = CX7/$CX$13)
        if bu in bus_validos:
            ws.cell(row=r, column=11).value = (
                f"=J{r}/SUM($J${bu_ini}:$J${fila_validos_fin})"
            )
            ws.cell(row=r, column=11).number_format = FMT_PCT
        else:
            ws.cell(row=r, column=11).value = ""  # Capex/MCS no tienen %PCT

    # Total (suma de TODOS incluidos Capex/MCS, igual que CV13)
    bu_fin = bu_ini + len(bus_orden) - 1
    r_total = bu_fin + 1
    ws.cell(row=r_total, column=9, value="Total").font = FONT_BOLD
    ws.cell(row=r_total, column=10).value = f"=SUM(J{bu_ini}:J{bu_fin})"
    ws.cell(row=r_total, column=10).number_format = FMT_MONTO
    ws.cell(row=r_total, column=10).fill = FILL_TOTAL
    ws.cell(row=r_total, column=10).font = FONT_BOLD

    # Total para %PCT (solo válidos)
    r_total_pct = r_total + 1
    ws.cell(row=r_total_pct, column=9, value="Total %PCT").font = Font(italic=True)
    ws.cell(row=r_total_pct, column=10).value = f"=SUM($J${bu_ini}:$J${fila_validos_fin})"
    ws.cell(row=r_total_pct, column=10).number_format = FMT_MONTO
    ws.cell(row=r_total_pct, column=10).font = Font(italic=True)

    widths = {1: 3, 2: 10, 3: 22, 4: 18, 5: 18, 6: 12, 7: 14, 8: 18, 9: 14, 10: 16, 11: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    return {
        "bu_range": f"$I${bu_ini}:$I${bu_fin}",
        "amount_range": f"$J${bu_ini}:$J${bu_fin}",
        "pct_range": f"$K${bu_ini}:$K${bu_fin}",
    }


# ════════════════════════════════════════════════════════════
# HOJA LAND
# ════════════════════════════════════════════════════════════
def _hoja_land(
    wb: Workbook,
    df: pd.DataFrame,
    tarifa: float,
    bus_summary: List[str],
    bus_excluir_pct: List[str] = None,
) -> Dict[str, str]:
    """
    Layout (réplica del Excel original):
      • Cell H2 = tarifa fija (réplica $BZ$1)
      • Cols B:G = Reference, BU, PN, Gross Weight, %Propot, Amount
      • Cols I:K = Resumen por BU (excluye Miscelaneus/Machine del %PCT)
    """
    if bus_excluir_pct is None:
        bus_excluir_pct = ["Miscelaneus", "Machine"]

    ws = wb.create_sheet("Land")

    # ───── Tarifa fija ─────
    ws["G2"] = "Tarifa Reference:"
    ws["G2"].font = FONT_BOLD
    ws["H2"] = tarifa
    ws["H2"].number_format = FMT_MONTO
    ws["H2"].fill = FILL_TARIFA
    ws["H2"].font = Font(bold=True, color="C00000")

    # ───── Encabezados ─────
    headers = [("B5", "Reference"), ("C5", "BU"), ("D5", "PN"),
               ("E5", "Gross Weight"), ("F5", "% Propot"), ("G5", "Amount")]
    for addr, val in headers:
        c = ws[addr]; c.value = val; c.font = FONT_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")

    # ───── Datos ─────
    col_ref = "Reference"
    col_bu = "BU Final" if "BU Final" in df.columns else "BU"
    col_item = "Item"
    col_peso = "Peso Bruto"

    fila_ini, n = 6, len(df)
    fila_fin = fila_ini + n - 1

    for i, (_, row) in enumerate(df.iterrows()):
        r = fila_ini + i
        ws.cell(row=r, column=2, value=str(row[col_ref]))
        ws.cell(row=r, column=3, value=str(row.get(col_bu, "")))
        ws.cell(row=r, column=4, value=str(row[col_item]))
        ws.cell(row=r, column=5, value=float(row[col_peso]) if pd.notna(row[col_peso]) else 0)
        ws.cell(row=r, column=5).number_format = FMT_PESO

        # 🔑 F = %Propot (réplica BY6)
        ws.cell(row=r, column=6).value = (
            f"=E{r}*1/SUMIFS($E${fila_ini}:$E${fila_fin},$B${fila_ini}:$B${fila_fin},B{r})"
        )
        ws.cell(row=r, column=6).number_format = FMT_PCT
        ws.cell(row=r, column=6).fill = FILL_FORMULA

        # 🔑 G = Amount (réplica BZ6 = BY6 * $BZ$1)
        ws.cell(row=r, column=7).value = f"=F{r}*$H$2"
        ws.cell(row=r, column=7).number_format = FMT_MONTO
        ws.cell(row=r, column=7).fill = FILL_FORMULA

    # ───── Resumen por BU (cols I:K) ─────
    ws["I5"] = "📊 Resumen por BU"
    ws["I5"].font = FONT_HDR; ws["I5"].fill = FILL_HDR
    ws.merge_cells("I5:K5")

    for col_letter, val in zip(["I", "J", "K"], ["BU", "Sum", "%PCT"]):
        c = ws[f"{col_letter}6"]; c.value = val; c.font = FONT_HDR; c.fill = FILL_HDR

    bus_en_df = sorted(df[col_bu].dropna().unique().tolist())
    bus_validos = [b for b in bus_en_df if b not in bus_excluir_pct]
    bus_excl_present = [b for b in bus_en_df if b in bus_excluir_pct]
    bus_orden = bus_validos + bus_excl_present

    bu_ini = 7
    fila_validos_fin = bu_ini + len(bus_validos) - 1

    for i, bu in enumerate(bus_orden):
        r = bu_ini + i
        ws.cell(row=r, column=9, value=bu)
        # 🔑 SUMIFS por BU (réplica CD6)
        ws.cell(row=r, column=10).value = (
            f"=SUMIFS($G${fila_ini}:$G${fila_fin},$C${fila_ini}:$C${fila_fin},I{r})"
        )
        ws.cell(row=r, column=10).number_format = FMT_MONTO

        # 🔑 %PCT EXCLUYE Misc/Machine (réplica CE6 = CD6/SUM($CD$6:$CD$9))
        if bu in bus_validos:
            ws.cell(row=r, column=11).value = (
                f"=J{r}/SUM($J${bu_ini}:$J${fila_validos_fin})"
            )
            ws.cell(row=r, column=11).number_format = FMT_PCT
        else:
            ws.cell(row=r, column=11).value = ""

    bu_fin = bu_ini + len(bus_orden) - 1
    r_total = bu_fin + 1
    ws.cell(row=r_total, column=9, value="Total").font = FONT_BOLD
    ws.cell(row=r_total, column=10).value = f"=SUM(J{bu_ini}:J{bu_fin})"
    ws.cell(row=r_total, column=10).number_format = FMT_MONTO
    ws.cell(row=r_total, column=10).fill = FILL_TOTAL
    ws.cell(row=r_total, column=10).font = FONT_BOLD

    widths = {1: 3, 2: 22, 3: 12, 4: 22, 5: 14, 6: 12, 7: 14, 8: 18, 9: 14, 10: 14, 11: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    return {
        "bu_range": f"$I${bu_ini}:$I${bu_fin}",
        "amount_range": f"$J${bu_ini}:$J${bu_fin}",
        "pct_range": f"$K${bu_ini}:$K${bu_fin}",
    }


# ════════════════════════════════════════════════════════════
# HOJA SUMMARY (XLOOKUP idéntico a tu Excel original)
# ════════════════════════════════════════════════════════════
def _hoja_summary(
    wb: Workbook,
    bus_summary: List[str],
    bus_ranges: Dict[str, Dict[str, str]],
) -> None:
    """
    Construye la hoja Summary IDÉNTICA al Excel original:
    
    Estructura (réplica de las filas B3:H12):
      Fila 3:  Type | M01 | M19 | M23 | M45 | M46
      Fila 4:  Sea %PCT       | XLOOKUP a Sea!pct_range
      Fila 5:  Land %PCT      | XLOOKUP a Land!pct_range
      Fila 6:  Outbound %PCT  | XLOOKUP a Outbound!pct_range
      Fila 9:  Viewer | Arg.Var$ | M01..M46
      Fila 10: Sea       | SUM(D10:H10) | XLOOKUP montos Sea
      Fila 11: Land      | SUM(D11:H11) | XLOOKUP montos Land
      Fila 12: Outbound  | SUM(D12:H12) | XLOOKUP montos Outbound
    """
    ws = wb.create_sheet("Summary", 0)  # Primera hoja

    # ───── Título ─────
    ws["B2"] = "📊 SUMMARY — Consolidación por BU"
    ws["B2"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("B2:H2")

    # ───── Sección 1: %PCT (filas 3-6) ─────
    ws["C3"] = "Type"
    ws["C3"].font = FONT_HDR; ws["C3"].fill = FILL_HDR
    ws["C3"].alignment = Alignment(horizontal="center")

    for i, bu in enumerate(bus_summary):
        col = get_column_letter(4 + i)  # D=4
        c = ws[f"{col}3"]; c.value = bu
        c.font = FONT_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")

    operaciones_pct = [
        ("Sea %PCT", "Sea", 4),
        ("Land %PCT", "Land", 5),
        ("Outbound %PCT", "Outbound", 6),
    ]

    for label, hoja, fila in operaciones_pct:
        ws[f"C{fila}"] = label
        ws[f"C{fila}"].font = FONT_BOLD

        if hoja not in bus_ranges:
            continue

        bu_rng = bus_ranges[hoja]["bu_range"]
        pct_rng = bus_ranges[hoja]["pct_range"]

        for i, bu in enumerate(bus_summary):
            col = get_column_letter(4 + i)
            # 🔑 XLOOKUP IDÉNTICO al Excel original
            ws[f"{col}{fila}"].value = (
                f"=IFERROR(XLOOKUP({col}$3,{hoja}!{bu_rng},{hoja}!{pct_rng},0),0)"
            )
            ws[f"{col}{fila}"].number_format = FMT_PCT
            ws[f"{col}{fila}"].fill = FILL_FORMULA

    # ───── Sección 2: Arg.Var $ (filas 9-12) ─────
    ws["B9"] = "Viewer"
    ws["B9"].font = FONT_HDR; ws["B9"].fill = FILL_HDR
    ws["C9"] = "Arg. Var $"
    ws["C9"].font = FONT_HDR; ws["C9"].fill = FILL_HDR

    for i, bu in enumerate(bus_summary):
        col = get_column_letter(4 + i)
        c = ws[f"{col}9"]; c.value = bu
        c.font = FONT_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")

    operaciones_monto = [
        ("Sea", 10),
        ("Land", 11),
        ("Outbound", 12),
    ]

    for hoja, fila in operaciones_monto:
        ws[f"B{fila}"] = hoja
        ws[f"B{fila}"].font = FONT_BOLD

        # Columna C = SUM(D:H) — réplica del Excel
        ws[f"C{fila}"].value = f"=SUM(D{fila}:H{fila})"
        ws[f"C{fila}"].number_format = FMT_MONTO
        ws[f"C{fila}"].font = FONT_BOLD
        ws[f"C{fila}"].fill = FILL_TOTAL

        if hoja not in bus_ranges:
            continue

        bu_rng = bus_ranges[hoja]["bu_range"]
        amt_rng = bus_ranges[hoja]["amount_range"]

        for i, bu in enumerate(bus_summary):
            col = get_column_letter(4 + i)
            # 🔑 XLOOKUP IDÉNTICO al Excel original
            ws[f"{col}{fila}"].value = (
                f"=IFERROR(XLOOKUP({col}$9,{hoja}!{bu_rng},{hoja}!{amt_rng},0),0)"
            )
            ws[f"{col}{fila}"].number_format = FMT_MONTO
            ws[f"{col}{fila}"].fill = FILL_FORMULA

    # ───── Anchos ─────
    widths = {1: 3, 2: 12, 3: 16, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Bordes a tabla %PCT
    for fila in range(3, 7):
        for col in range(3, 4 + len(bus_summary)):
            ws.cell(row=fila, column=col).border = BORDER

    for fila in range(9, 13):
        for col in range(2, 4 + len(bus_summary)):
            ws.cell(row=fila, column=col).border = BORDER


# ════════════════════════════════════════════════════════════
# 🎯 FUNCIÓN PRINCIPAL
# ════════════════════════════════════════════════════════════
def generar_excel_completo(
    df_outbound: Optional[pd.DataFrame] = None,
    df_tarifas_outbound: Optional[pd.DataFrame] = None,
    df_sea: Optional[pd.DataFrame] = None,
    df_land: Optional[pd.DataFrame] = None,
    tarifa_outbound: float = 1500.0,
    tarifa_sea: float = 2500.0,
    tarifa_land: float = 1200.0,
    bus_summary: List[str] = None,
    bus_excluir_sea: List[str] = None,
    bus_excluir_land: List[str] = None,
) -> BytesIO:
    """
    Genera un archivo Excel COMPLETO con fórmulas vivas idénticas al Excel original.

    Args:
        df_outbound:          DataFrame ya procesado (con BU Final)
        df_tarifas_outbound:  Tabla [Reference, Fix Cost] (None = usar default)
        df_sea, df_land:      DataFrames ya procesados
        tarifa_*:             Tarifas default (overrideables)
        bus_summary:          BUs a mostrar en Summary (default M01..M46)
        bus_excluir_sea:      BUs sin %PCT en Sea (default ['Capex','MCS'])
        bus_excluir_land:     BUs sin %PCT en Land (default ['Miscelaneus','Machine'])

    Returns:
        BytesIO con el XLSX (listo para st.download_button)
    """
    if bus_summary is None:
        bus_summary = BUS_SUMMARY_DEFAULT.copy()
    if bus_excluir_sea is None:
        bus_excluir_sea = ["Capex", "MCS"]
    if bus_excluir_land is None:
        bus_excluir_land = ["Miscelaneus", "Machine"]

    wb = Workbook()
    wb.remove(wb.active)  # Quitar hoja default

    bus_ranges: Dict[str, Dict[str, str]] = {}

    # Construir hojas en orden (Summary se crea al final pero se inserta en pos 0)
    if df_sea is not None and len(df_sea) > 0:
        bus_ranges["Sea"] = _hoja_sea(wb, df_sea, tarifa_sea, bus_summary, bus_excluir_sea)

    if df_land is not None and len(df_land) > 0:
        bus_ranges["Land"] = _hoja_land(wb, df_land, tarifa_land, bus_summary, bus_excluir_land)

    if df_outbound is not None and len(df_outbound) > 0:
        bus_ranges["Outbound"] = _hoja_outbound(
            wb, df_outbound, df_tarifas_outbound, tarifa_outbound, bus_summary
        )

    # Summary AL FINAL (necesita los rangos) pero se inserta primero
    if bus_ranges:
        _hoja_summary(wb, bus_summary, bus_ranges)

    # Serializar
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
