"""
═══════════════════════════════════════════════════════════════
DETECTOR AUTOMÁTICO DE HOJA POR CONTENIDO
═══════════════════════════════════════════════════════════════
Analiza cada hoja del archivo y elige la más probable de ser
la hoja de datos para la operación solicitada, sin depender del
nombre de la hoja.

🔧 v2 — Fixes:
  1. score se aplana correctamente (extrae int de dict)
  2. criticas se obtiene de COLUMNAS_POR_OPERACION con .get() seguro
  3. Logging mejorado para diagnóstico
═══════════════════════════════════════════════════════════════
"""
import pandas as pd
from typing import Optional, Dict, List, Tuple
from openpyxl import load_workbook

from src.ingesta.detector_header import detectar_fila_encabezado
from src.ingesta.mapeo_columnas import (
    mapear_columnas_dataframe,
    calcular_score_hoja,
    COLUMNAS_POR_OPERACION,
)
from src.utils.logger import configurar_logger

logger = configurar_logger("detector_hojas")


# Hojas a IGNORAR siempre (son metadata o no contienen datos)
HOJAS_IGNORADAS = [
    "xdo_metadata",
    "metadata",
    "config",
    "configuracion",
    "instructions",
    "instrucciones",
    "readme",
    "info",
    "summary",  # Es el reporte final, no datos crudos
    "consolidado",
]


def listar_hojas(archivo) -> List[str]:
    """Lista todas las hojas del archivo Excel."""
    try:
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        wb = load_workbook(archivo, read_only=True, data_only=True)
        hojas = wb.sheetnames
        wb.close()
        return hojas
    except Exception as e:
        logger.error(f"❌ No se pudo leer el archivo: {e}")
        return []


def analizar_hoja(
    archivo,
    nombre_hoja: str,
    operacion: str,
) -> Dict:
    """
    Analiza una hoja específica y devuelve su información:
        - fila_header: dónde están los encabezados
        - score: qué tan probable es que sea de esta operación (INT)
        - columnas_logicas: qué columnas críticas/opcionales tiene
    """
    resultado = {
        "hoja": nombre_hoja,
        "fila_header": None,
        "num_filas": 0,
        "num_columnas": 0,
        "score": 0,
        "columnas_logicas": [],
        "columnas_criticas_presentes": [],
        "columnas_criticas_faltantes": [],
        "ignorada": False,
        "motivo_ignorada": None,
    }

    # 1. Verificar si debe ignorarse por nombre
    nombre_lower = nombre_hoja.lower().strip()
    for ignorada in HOJAS_IGNORADAS:
        if ignorada in nombre_lower:
            resultado["ignorada"] = True
            resultado["motivo_ignorada"] = f"Nombre contiene '{ignorada}'"
            return resultado

    # 2. Detectar fila de encabezado
    fila_header, num_cols_reconocidas = detectar_fila_encabezado(archivo, nombre_hoja)

    if fila_header is None:
        resultado["motivo_ignorada"] = "No se encontró encabezado válido"
        resultado["ignorada"] = True
        return resultado

    resultado["fila_header"] = fila_header

    # 3. Leer la hoja con el encabezado detectado
    try:
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        df = pd.read_excel(
            archivo,
            sheet_name=nombre_hoja,
            header=fila_header,
            dtype=object,
        )
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        resultado["motivo_ignorada"] = f"Error al leer: {e}"
        resultado["ignorada"] = True
        return resultado

    resultado["num_filas"] = len(df)
    resultado["num_columnas"] = len(df.columns)

    # 4. Mapear columnas con alias
    df_mapeado, mapeo, no_mapeadas = mapear_columnas_dataframe(df, operacion)
    columnas_logicas = list(set(mapeo.values()))
    resultado["columnas_logicas"] = columnas_logicas

    # ═══════════════════════════════════════════════════════
    # 🔧 FIX 1: calcular_score_hoja devuelve un DICT
    #          → debemos extraer el campo 'score' como INT
    # ═══════════════════════════════════════════════════════
    score_info = calcular_score_hoja(df_mapeado, operacion)
    
    # Manejar ambos casos: si devuelve dict o si devuelve int directo
    if isinstance(score_info, dict):
        resultado["score"] = int(score_info.get("score", 0))
        # Aprovechar info adicional si está disponible
        if "presentes" in score_info:
            resultado["columnas_criticas_presentes"] = list(score_info["presentes"])
        if "faltantes" in score_info:
            resultado["columnas_criticas_faltantes"] = list(score_info["faltantes"])
    else:
        # Fallback: si por algún motivo devuelve int directo
        resultado["score"] = int(score_info or 0)

    # ═══════════════════════════════════════════════════════
    # 🔧 FIX 2: Identificar críticas presentes/faltantes
    #          Si no se obtuvieron del score_info, calcularlas aquí
    # ═══════════════════════════════════════════════════════
    config = COLUMNAS_POR_OPERACION.get(operacion, {})
    
    # Soportar formato dict (nuevo) y lista (legacy)
    if isinstance(config, dict):
        criticas = config.get("criticas", [])
    elif isinstance(config, list):
        criticas = config
    else:
        criticas = []
    
    # Solo llenar si no estaban ya en score_info
    if not resultado["columnas_criticas_presentes"]:
        resultado["columnas_criticas_presentes"] = [
            c for c in criticas if c in columnas_logicas
        ]
    if not resultado["columnas_criticas_faltantes"]:
        resultado["columnas_criticas_faltantes"] = [
            c for c in criticas if c not in columnas_logicas
        ]

    return resultado


def detectar_hoja_optima(
    archivo,
    operacion: str,
    score_minimo: int = 25,
) -> Tuple[Optional[Dict], List[Dict]]:
    """
    Encuentra la hoja óptima para la operación dada.

    Args:
        archivo: Path o file-like del Excel
        operacion: 'land', 'outbound' o 'sea'
        score_minimo: Score mínimo aceptable (int)

    Returns:
        Tupla:
          - mejor_hoja: dict con info de la hoja elegida, o None
          - todas_las_hojas: lista de análisis de TODAS las hojas
    """
    logger.info("=" * 60)
    logger.info(f"🔍 DETECTANDO HOJA ÓPTIMA para operación '{operacion}'")
    logger.info("=" * 60)

    hojas = listar_hojas(archivo)
    if not hojas:
        return None, []

    logger.info(f"📚 Hojas encontradas: {hojas}")

    analisis = []
    for nombre_hoja in hojas:
        resultado = analizar_hoja(archivo, nombre_hoja, operacion)
        analisis.append(resultado)

        if resultado["ignorada"]:
            logger.info(
                f"   ⏭️ '{nombre_hoja}' ignorada: {resultado['motivo_ignorada']}"
            )
        else:
            logger.info(
                f"   📊 '{nombre_hoja}': score={resultado['score']} | "
                f"filas={resultado['num_filas']} | "
                f"críticas: {resultado['columnas_criticas_presentes']} | "
                f"faltantes: {resultado['columnas_criticas_faltantes']}"
            )

    # ═══════════════════════════════════════════════════════
    # 🔧 FIX 3: Filtrar candidatas — score es INT garantizado
    # ═══════════════════════════════════════════════════════
    candidatas = [
        a for a in analisis
        if not a["ignorada"] and int(a["score"]) >= score_minimo
    ]

    if not candidatas:
        logger.warning(
            f"⚠️ Ninguna hoja alcanzó el score mínimo ({score_minimo}). "
            f"Se requerirá selección manual."
        )
        return None, analisis

    # Ordenar por score descendente (score es int garantizado)
    candidatas.sort(key=lambda x: (-int(x["score"]), -int(x["num_filas"])))
    mejor = candidatas[0]

    logger.info(f"✅ HOJA ELEGIDA: '{mejor['hoja']}' (score: {mejor['score']})")
    logger.info(f"   Encabezado en fila: {mejor['fila_header'] + 1}")
    logger.info(f"   Columnas mapeadas: {mejor['columnas_logicas']}")

    return mejor, analisis