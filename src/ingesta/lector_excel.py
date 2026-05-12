"""
Lector de archivos Excel.
Implementa funciones independientes para cargar SEA, LAND y OUTBOUND.

IMPORTANTE (fix LAND): El archivo land tiene DOS columnas llamadas 'BU' 
(una en E con los BUs reales, otra en BV con Machine/Miscelaneus).
Por eso LAND se lee por POSICIÓN de columna (E, AZ, BC), no por nombre.
"""
import pandas as pd
from pathlib import Path
from typing import Union, BinaryIO
from openpyxl import load_workbook

from src.ingesta.excepciones import (
    ArchivoInvalidoError,
    HojaNoEncontradaError,
    DatosVaciosError,
)
from src.ingesta.normalizador import (
    normalizar_texto,
    normalizar_numerico,
    eliminar_filas_vacias,
    eliminar_filas_totales,
)
from src.ingesta.validador_columnas import validar_columnas
from src.utils.logger import configurar_logger

logger = configurar_logger("ingesta")


# ============================================================
# CONFIGURACIÓN POR TIPO DE OPERACIÓN
# ============================================================
CONFIG_HOJAS = {
    "sea": {
        "nombres_hoja": ["Sea", "SEA", "sea", "China Maritimos", "Maritimos"],
        "fila_encabezado": 5,
        "columnas_clave": ["BU", "Container Number", "Total Gross Weight"],
    },
    "land": {
        "nombres_hoja": ["land", "Land", "LAND", "Terrestre"],
        "fila_encabezado": 5,
        "columnas_clave": ["Reference", "BU", "Peso Bruto (Kgs)"],
    },
    "outbound": {
        "nombres_hoja": ["Outbound", "OUTBOUND", "outbound", "Exportaciones"],
        "fila_encabezado": 8,
        "columnas_clave": ["Reference", "Waybill Number", "Gross Weight"],
    },
}


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================
def _detectar_hoja(archivo, nombres_posibles: list) -> str:
    """Busca el nombre exacto de la hoja en el archivo."""
    try:
        # Si es un objeto file-like (Streamlit), reset al inicio
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        wb = load_workbook(archivo, read_only=True, data_only=True)
        hojas_disponibles = wb.sheetnames
        wb.close()
    except Exception as e:
        raise ArchivoInvalidoError(f"No se pudo abrir el archivo: {e}")
    
    for nombre in nombres_posibles:
        if nombre in hojas_disponibles:
            return nombre
    
    hojas_lower = {h.lower(): h for h in hojas_disponibles}
    for nombre in nombres_posibles:
        if nombre.lower() in hojas_lower:
            return hojas_lower[nombre.lower()]
    
    raise HojaNoEncontradaError(nombres_posibles[0], hojas_disponibles)


def _leer_excel_con_header(archivo, hoja: str, fila_encabezado: int) -> pd.DataFrame:
    """Lee un Excel especificando la fila de encabezado."""
    try:
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        df = pd.read_excel(
            archivo,
            sheet_name=hoja,
            header=fila_encabezado - 1,
            dtype=object,
        )
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception as e:
        raise ArchivoInvalidoError(f"Error al leer la hoja '{hoja}': {e}")


def _leer_excel_por_posicion(archivo, hoja: str, fila_encabezado: int, 
                              columnas_excel: list, nombres_finales: list) -> pd.DataFrame:
    """
    🆕 NUEVA FUNCIÓN: Lee un Excel seleccionando columnas por POSICIÓN (letra),
    no por nombre. Útil cuando hay columnas duplicadas con el mismo nombre.
    
    Args:
        archivo: Ruta o file-like
        hoja: Nombre de la hoja
        fila_encabezado: Fila donde están los encabezados (1-indexed)
        columnas_excel: Lista de letras de columna ['B', 'C', 'D', 'E', 'AZ', 'BC']
        nombres_finales: Nombres a asignar a esas columnas
    
    Returns:
        DataFrame con solo las columnas solicitadas, con nombres correctos
    """
    try:
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        
        # Construir string "B,C,D,E,AZ,BC" para usecols
        usecols_str = ",".join(columnas_excel)
        
        df = pd.read_excel(
            archivo,
            sheet_name=hoja,
            header=fila_encabezado - 1,
            usecols=usecols_str,
            dtype=object,
        )
        
        # Renombrar columnas por POSICIÓN (orden de columnas_excel)
        # pandas las lee en el orden alfabético de Excel: B, C, D, E, AZ, BC
        # Necesitamos ordenar nuestros nombres_finales según ese orden
        from openpyxl.utils import column_index_from_string
        
        # Crear pares (letra, nombre) y ordenar por índice de columna
        pares = list(zip(columnas_excel, nombres_finales))
        pares.sort(key=lambda x: column_index_from_string(x[0]))
        
        # Renombrar en el orden correcto
        nombres_ordenados = [nombre for letra, nombre in pares]
        df.columns = nombres_ordenados
        
        return df
    
    except Exception as e:
        raise ArchivoInvalidoError(f"Error al leer la hoja '{hoja}' por posición: {e}")


# ============================================================
# FUNCIÓN: CARGAR SEA
# ============================================================
def cargar_sea(archivo) -> pd.DataFrame:
    """Carga el reporte SEA."""
    logger.info("📥 Cargando archivo SEA...")
    config = CONFIG_HOJAS["sea"]
    
    hoja = _detectar_hoja(archivo, config["nombres_hoja"])
    logger.info(f"   Hoja detectada: '{hoja}'")
    
    df = _leer_excel_con_header(archivo, hoja, config["fila_encabezado"])
    logger.info(f"   Filas leídas: {len(df)} | Columnas: {len(df.columns)}")
    
    # Manejar columnas duplicadas: si hay dos 'BU', tomar la primera
    df = df.loc[:, ~df.columns.duplicated(keep="first")]
    
    validar_columnas(df, "sea")
    
    df["BU"] = normalizar_texto(df["BU"])
    df["Item Code"] = normalizar_texto(df["Item Code"])
    df["Container Number"] = normalizar_texto(df["Container Number"])
    df["Total Gross Weight"] = normalizar_numerico(df["Total Gross Weight"])
    
    df = eliminar_filas_vacias(df, ["Container Number", "Item Code"])
    df = eliminar_filas_totales(df, "BU")
    
    if len(df) == 0:
        raise DatosVaciosError("El archivo SEA no contiene datos válidos.")
    
    logger.info(f"✅ SEA cargado correctamente: {len(df)} registros válidos")
    return df


# ============================================================
# FUNCIÓN: CARGAR LAND (🔧 CORREGIDA)
# ============================================================
def cargar_land(archivo) -> pd.DataFrame:
    """
    Carga el reporte LAND.
    
    🔧 FIX: El archivo tiene DOS columnas 'BU' (una real en E, otra en BV con 
    Machine/Miscelaneus del proceso). Por eso leemos por POSICIÓN de columna:
    
    Estructura esperada:
        - Hoja: 'land' (o variantes)
        - Encabezado: fila 5
        - Columnas a leer (por letra):
            B  → Inbound/Outbound
            C  → Method
            D  → Reference
            E  → BU              ← BU REAL (M19, M23, M45, M46, etc.)
            AZ → Peso Bruto (Kgs)
            BC → No. Parte Prov.
    """
    logger.info("📥 Cargando archivo LAND...")
    config = CONFIG_HOJAS["land"]
    
    # 1. Detectar hoja
    hoja = _detectar_hoja(archivo, config["nombres_hoja"])
    logger.info(f"   Hoja detectada: '{hoja}'")
    
    # 2. Leer SOLO las columnas correctas por posición (evita conflicto con BV)
    columnas_excel = ["B", "C", "D", "E", "AZ", "BC"]
    nombres_finales = [
        "Inbound/Outbound",
        "Method",
        "Reference",
        "BU",
        "Peso Bruto (Kgs)",
        "No. Parte Prov.",
    ]
    
    df = _leer_excel_por_posicion(
        archivo, hoja, config["fila_encabezado"],
        columnas_excel, nombres_finales
    )
    logger.info(f"   Filas leídas: {len(df)} | Columnas: {len(df.columns)}")
    logger.info(f"   Columnas asignadas: {list(df.columns)}")
    
    # 3. Validar columnas
    validar_columnas(df, "land")
    
    # 4. Normalización
    df["Inbound/Outbound"] = normalizar_texto(df["Inbound/Outbound"])
    df["Method"] = normalizar_texto(df["Method"])
    df["Reference"] = normalizar_texto(df["Reference"])
    df["BU"] = normalizar_texto(df["BU"])
    df["Peso Bruto (Kgs)"] = normalizar_numerico(df["Peso Bruto (Kgs)"])
    df["No. Parte Prov."] = normalizar_texto(df["No. Parte Prov."])
    
    # 5. Limpieza
    df = eliminar_filas_vacias(df, ["Reference", "No. Parte Prov."])
    df = eliminar_filas_totales(df, "Reference")
    
    if len(df) == 0:
        raise DatosVaciosError("El archivo LAND no contiene datos válidos.")
    
    # 6. Log de validación: mostrar BUs detectados
    bus_detectados = sorted(df["BU"].dropna().unique().tolist())
    logger.info(f"   ✅ BUs detectados en LAND: {bus_detectados}")
    
    logger.info(f"✅ LAND cargado correctamente: {len(df)} registros válidos")
    return df


def cargar_outbound(archivo) -> tuple:
    """
    Carga el reporte OUTBOUND junto con su tabla de costos variables.
    
    🔧 ESTRUCTURA REAL:
        Sección Extraction (datos): columnas BG-BO, fila 8
        Tabla de costos:            columnas BC-BE, fila 8
    
    Returns:
        tuple: (df_datos, df_costos)
            - df_datos: DataFrame con los items
            - df_costos: DataFrame con [Reference, Cost (USD), Fix Cost]
              o None si no se encuentra la tabla
    """
    logger.info("📥 Cargando archivo OUTBOUND...")
    config = CONFIG_HOJAS["outbound"]
    
    hoja = _detectar_hoja(archivo, config["nombres_hoja"])
    logger.info(f"   Hoja detectada: '{hoja}'")
    
    # ────────────────────────────────────────────────────────
    # 1. CARGAR LA SECCIÓN EXTRACTION (DATOS)
    # ────────────────────────────────────────────────────────
    columnas_excel = ["BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO"]
    nombres_finales = [
        "Inbound/Outbound", "Method", "Reference",
        "Customer", "ADDRESS", "BU",
        "Item", "Qty Pzas", "Gross Weight",
    ]
    
    fila_encabezado = config["fila_encabezado"]  # 8
    
    try:
        df = _leer_excel_por_posicion(
            archivo, hoja, fila_encabezado,
            columnas_excel, nombres_finales
        )
        logger.info(f"   Filas leídas (Extraction): {len(df)}")
    except Exception as e:
        logger.warning(f"Lectura por posición falló: {e}")
        df = _leer_excel_con_header(archivo, hoja, fila_encabezado)
        df = df.loc[:, ~df.columns.duplicated(keep="last")]
    
    validar_columnas(df, "outbound")
    
    # Normalización
    df["Inbound/Outbound"] = normalizar_texto(df["Inbound/Outbound"])
    df["Method"] = normalizar_texto(df["Method"])
    df["Reference"] = normalizar_texto(df["Reference"])
    df["BU"] = normalizar_texto(df["BU"])
    df["Gross Weight"] = normalizar_numerico(df["Gross Weight"])
    df["Item"] = normalizar_texto(df["Item"])
    df["Qty Pzas"] = normalizar_numerico(df["Qty Pzas"])
    
    if "Customer" in df.columns:
        df["Customer"] = normalizar_texto(df["Customer"])
    if "ADDRESS" in df.columns:
        df["ADDRESS"] = normalizar_texto(df["ADDRESS"])
    
    if "Waybill Number" not in df.columns:
        df["Waybill Number"] = df["Reference"]
    
    df = eliminar_filas_vacias(df, ["Reference", "Item"])
    df = eliminar_filas_totales(df, "Reference")
    
    if len(df) == 0:
        raise DatosVaciosError("El archivo OUTBOUND no contiene datos válidos.")
    
    # ────────────────────────────────────────────────────────
    # 2. 🆕 CARGAR LA TABLA DE COSTOS VARIABLES (BC:BE, fila 8)
    # ────────────────────────────────────────────────────────
    df_costos = _cargar_tabla_costos_outbound(archivo, hoja, fila_encabezado)
    
    if df_costos is not None and len(df_costos) > 0:
        logger.info(f"   ✅ Tabla de costos cargada: {len(df_costos)} references")
        total_fix_cost = df_costos["Fix Cost"].sum()
        logger.info(f"   💰 Suma total Fix Cost: ${total_fix_cost:,.2f}")
    else:
        logger.warning("   ⚠️ No se encontró tabla de costos. Se usará costo fijo.")
    
    # Logs finales
    bus_directos = df["BU"].dropna().unique().tolist()
    refs_unicos = df["Reference"].nunique()
    logger.info(f"   ✅ References únicas: {refs_unicos}")
    logger.info(f"   ✅ BUs: {sorted(bus_directos)}")
    logger.info(f"✅ OUTBOUND cargado: {len(df)} registros válidos")
    
    return df, df_costos


def _cargar_tabla_costos_outbound(archivo, hoja: str, fila_encabezado: int):
    """
    🆕 Carga la tabla de costos variables del Outbound (columnas BC:BE).
    
    Estructura esperada (fila 8):
        BC: Reference
        BD: Cost (USD)        ← costo consumido (puede ser $0)
        BE: Fix Cost          ← costo asignado/presupuestado ✅
    
    Returns:
        DataFrame con [Reference, Cost (USD), Fix Cost] o None
    """
    try:
        df_costos = _leer_excel_por_posicion(
            archivo, hoja, fila_encabezado,
            columnas_excel=["BC", "BD", "BE"],
            nombres_finales=["Reference", "Cost (USD)", "Fix Cost"],
        )
        
        # Limpieza
        df_costos["Reference"] = normalizar_texto(df_costos["Reference"])
        df_costos["Cost (USD)"] = normalizar_numerico(df_costos["Cost (USD)"])
        df_costos["Fix Cost"] = normalizar_numerico(df_costos["Fix Cost"])
        
        # Eliminar filas vacías o con Reference inválida
        df_costos = df_costos.dropna(subset=["Reference"])
        df_costos = df_costos[df_costos["Reference"].str.strip() != ""]
        df_costos = df_costos[df_costos["Fix Cost"] > 0]
        df_costos = df_costos.reset_index(drop=True)
        
        return df_costos if len(df_costos) > 0 else None
    
    except Exception as e:
        logger.warning(f"No se pudo cargar tabla de costos: {e}")
        return None
    
# ============================================================
# METADATA
# ============================================================
def obtener_metadata(archivo) -> dict:
    """Retorna información general del archivo Excel."""
    try:
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        wb = load_workbook(archivo, read_only=True, data_only=True)
        metadata = {
            "hojas": wb.sheetnames,
            "total_hojas": len(wb.sheetnames),
        }
        wb.close()
        return metadata
    except Exception as e:
        raise ArchivoInvalidoError(f"No se pudo leer metadata: {e}")
